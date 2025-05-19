from openpyxl import load_workbook
from datetime import datetime
from pathlib import Path

def fill_po_template(template_path, matched_products, customer_info):
    """
    Fill the main_template.xlsx with the matched products and customer information using openpyxl.
    """
    wb = load_workbook(template_path)
    ws = wb.active

    # PO Number (F4)
    po_number = f"PO-{customer_info['Unidades Maduras']}-{datetime.now().strftime('%Y%m%d')}"
    ws['F4'] = po_number

    # Date (E7)
    ws['E7'] = datetime.now().strftime('%d/%m/%Y')

    # Customer info (Enviar a: E10, E11, E12, E13)
    # We'll use: Name, Address, Contact, and maybe a second line of address or contact if available
    ws['E10'] = customer_info.get('Unidades Maduras', '')
    ws['E11'] = customer_info.get('Domicilios', '')
    ws['E12'] = customer_info.get('Contacto Compras', '')
    ws['E13'] = customer_info.get('Contacto Compras2', '')

    # Product table starts at row 18 (B18:G18), headers are in row 17
    start_row = 18
    for i, product in enumerate(matched_products):
        row = start_row + i
        ws[f'B{row}'] = product['matched_name'] or ''  # Descripción
        ws[f'C{row}'] = product['matched_code'] or ''  # Código
        ws[f'D{row}'] = product['quantity'] if product['quantity'] is not None else ''  # Cantidad
        ws[f'E{row}'] = product['matched_unit'] or ''  # Unidad
        ws[f'F{row}'] = product['unit_price'] if product['unit_price'] is not None else ''  # Precio Unitario (MXN)
        if product['quantity'] is not None and product['unit_price'] is not None:
            ws[f'G{row}'] = product['quantity'] * product['unit_price']  # Total
        else:
            ws[f'G{row}'] = ''

    # Save to a new file in memory
    output_dir = Path("generated_pos")
    output_dir.mkdir(exist_ok=True)
    output_path = output_dir / f"po_{po_number}.xlsx"
    wb.save(output_path)
    return output_path, po_number 